import csv
import pandas as pd
import re
from pathlib import Path
from typing import Dict, List, Optional
from collections import defaultdict
from openpyxl import load_workbook
from src.survey.helper import *
from src.survey.constants import *

        
def prepare_files_descriptions_data(json_file_path: str) -> dict:
    """
    Prepares data descriptions for files from a JSON file.

    Args:
        json_file_path (str): The path of the JSON file.

    Returns:
        dict: The prepared data descriptions for files.
    """
    main_json_data = json_reader(json_file_path)["codeBook"]
    files_descr_values = main_json_data["fileDscr"]
    cols_data_descr = main_json_data["dataDscr"]["var"]
    interested_cols = ["@ID", "@files", "@name", "labl", "varFormat"]
    cols_filtered_dicts = filter_dicts_by_keys(cols_data_descr, interested_cols)
    
    data_files_descriptions = {}
    if cols_filtered_dicts:
        for filtered_value in cols_filtered_dicts:
            file_key = str(filtered_value["@files"]).strip()
            col_name = str(filtered_value["@name"]).strip()

            label = filtered_value.get("labl", col_name)
            updating_value = {
                col_name: {
                    "context": label,
                    "type": filtered_value["varFormat"]["@type"],
                }
            }
            observation_level = {"ea": "no", "hhid": "no", "indiv": "no", "entid": "no"}

            levels_checker = False
            if col_name in observation_level:
                if col_name == "ea":
                    observation_level["ea"] = "yes"
                elif col_name == "hhid":
                    observation_level["hhid"] = "yes"
                elif col_name == "indiv":
                    observation_level["indiv"] = "yes"
                else:
                    observation_level["entid"] = "yes"
                levels_checker = True

            if file_key in data_files_descriptions:
                data_files_descriptions[file_key]["columns"].update(updating_value)

                if levels_checker:
                    data_files_descriptions[file_key]["levels"][
                        col_name
                    ] = observation_level[col_name]
            else:
                data_files_descriptions[file_key] = {
                    "columns": updating_value,
                    "levels": observation_level,
                }

        for value in files_descr_values:
            file_text = value["fileTxt"]
            if value["@ID"] in data_files_descriptions:
                filename = file_text["fileName"].removesuffix(".dta")
                data_files_descriptions[filename] = data_files_descriptions.pop(
                    value["@ID"]
                )
                data_files_descriptions[filename]["description"] = file_text["fileCont"]

        data_files_descriptions = sortable_key_value_dictionary(data_files_descriptions)

    return data_files_descriptions


def write_dict_to_file(data, output_path, filename, file_type="xlsx"):
    """
    Writes a dictionary to a file in the specified format.

    Args:
        data: The dictionary to write to the file.
        output_path (str): The path of the output directory.
        filename (str): The name of the output file.
        file_type (str, optional): The file format to use ("csv", "xlsx", or "json"). Defaults to "xlsx".

    Returns:
        None
    """
    path = Path(f"{output_path}/{filename}.{file_type}")

    if not path.exists():
        container = []

        for file in data:
            builder = {"filename": file}

            for key, value in data[file].items():
                if isinstance(value, dict):
                    value = (
                        {"number_of_columns": len(value)} if key == "columns" else value
                    )
                    builder.update(value)
                else:
                    builder[key] = value

            container.append(builder)
            builder = {}

        output_dir = Path(output_path)
        output_dir.mkdir(parents=True, exist_ok=True)

        file_path = output_dir / (filename + "." + file_type)
        if file_type == "csv":
            with open(file_path, "w", newline="") as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=container[0].keys())
                writer.writeheader()
                writer.writerows(container)
        elif file_type == "xlsx":
            df = pd.DataFrame(container)
            df.to_excel(file_path, sheet_name="File Descriptions", index=False)
        elif file_type == "json":
            with open(file_path, "w") as jsonfile:
                json.dump(container, jsonfile, indent=4)
        else:
            print("Invalid file type. Supported types are 'csv', 'xlsx', and 'json'.")

        print("DONE! Finished creating files description data dictionary")
    else:
        print("DONE! File already exists.")


def add_extract_dict_to_excel(
    workbook_path: str,
    data: dict,
    excluding_columns: list,
    sheet_name: str = "Indicator-Domain - Indicator",
    column1: str = "Indicator_Domain",
    column2: str = "Indicator",
) -> tuple:
    """
    Adds a dictionary to an existing Excel workbook in the specified sheet.

    Args:
        workbook_path (str): The path of the Excel workbook.
        data (dict): The dictionary to add to the workbook.
        excluding_columns (list): The list of columns to exclude.
        sheet_name (str, optional): The name of the sheet to add the data. Defaults to "Indicator-Domain - Indicator".
        column1 (str, optional): The column header for Indicator_Domain. Defaults to "Indicator_Domain".
        column2 (str, optional): The column header for Indicator. Defaults to "Indicator".

    Returns:
        tuple: A tuple containing the indicator_domain_mapping, columns_filename_lookup, and file_columns_question_lookup.
    """
    workbook = load_workbook(workbook_path)

    if workbook.active:
        (
            file_columns_question_lookup,
            indicator_domain_mapping,
            columns_filename_lookup,
        ) = ({}, {}, {})
        filenames_indicator = {}

        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet("Dropdown Constants")
            sheet = workbook.create_sheet(sheet_name)

            # Write the column header
            sheet["A1"] = "Filenames"
            sheet["B1"] = "Features"
            sheet["C1"] = "Question/Context"
            sheet["D1"] = column1
            sheet["E1"] = column2

            row_adder = 0
            for filename in data:
                columns = data[filename]["columns"]

                # Write the keys as rows under the "Features" column
                for row_num, key in enumerate(columns.keys(), start=2):
                    row_num += row_adder
                    survey_question = columns[key]["context"]
                    sheet[f"A{row_num}"] = filename
                    sheet[f"B{row_num}"] = key
                    sheet[f"C{row_num}"] = survey_question

                    # Making the questionnaire lookup
                    if key not in excluding_columns:
                        file_columns_question_lookup[
                            f"{filename}_{key}"
                        ] = survey_question
                row_adder = row_num + 1
            workbook.save(workbook_path)
            print("DONE! Worksheets and lookups have been populated.")
        else:
            # TODO Extract out this for resuability elsewhere
            sheet = workbook[sheet_name]

            for row in sheet.iter_rows(min_row=2):
                values = tuple([cell.value for cell in row])

                assert len(values) == 5

                filename, column, survey_question, domain, indicator = values
                if filename and "phx" not in filename:
                    # Populating lookup values
                    if column not in excluding_columns:
                        file_columns_question_lookup[
                            f"{filename}_{column}"
                        ] = survey_question

                    if domain and indicator:
                        # Populating container's values
                        if domain in indicator_domain_mapping:
                            if indicator in indicator_domain_mapping[domain]:
                                indicator_domain_mapping[domain][indicator].add(column)
                            else:
                                indicator_domain_mapping[domain][indicator] = set(
                                    [column]
                                )
                        else:
                            indicator_domain_mapping[domain] = {
                                indicator: set([column])
                            }

                        # if domain == "Demographic and household data":
                        if filename in filenames_indicator:
                            filenames_indicator[filename].add(column)
                        else:
                            filenames_indicator[filename] = set([column])

                        # Populating lookup values
                        if column in columns_filename_lookup:
                            columns_filename_lookup[column].add(filename)
                        else:
                            columns_filename_lookup[column] = set([filename])

            print(f"DONE! {sheet_name} already exists and is ready for lookups.")

        return (
            indicator_domain_mapping,
            columns_filename_lookup,
            file_columns_question_lookup,
            filenames_indicator,
        )
    else:
        print("DONE! The data dictionary workbook path does not exist.")


def get_all_harmonized_files_dictionary(
    folder_path: str, file_extension_choices: List[str]
) -> Dict[str, str]:
    """
    Retrieves a dictionary containing all harmonized files in a folder.

    Args:
        folder_path (str): The path to the folder containing the files.
        file_extension_choices (List[str]): The list of file extensions to consider.

    Returns:
        Dict[str, str]: A dictionary mapping file names to their full paths.
    """

    def add_file_to_dict(file_path: Path, files_dict: Dict[str, str]) -> None:
        """
        Adds a file to the files dictionary.

        Args:
            file_path (Path): The path of the file to add.
            files_dict (Dict[str, str]): The dictionary to add the file to.

        Returns:
            None
        """

        if file_path.is_file():
            file_extension = file_path.suffix.lower()

            if file_extension in file_extension_choices:
                files_dict[file_path.stem] = str(file_path)

    folder_path = Path(folder_path)
    files_dict = {}

    for path in folder_path.rglob("*"):
        add_file_to_dict(path, files_dict)
    return sortable_key_value_dictionary(files_dict)


def preprocessing_transformer(
    df: pd.DataFrame, primary_columns: List[str], columns_to_check: List[str] = None
) -> pd.DataFrame:
    """
    Preprocesses the given dataframe by dropping duplicate rows, dropping NaN values, extracting year from 'wave' column,
    and converting selected columns to specific data types.

    Parameters:
        df (pd.DataFrame): The input dataframe.
        primary_columns (List[str]): A list of primary columns to be used for preprocessing.
        columns_to_check (List[str], optional): A list of columns to check for NaN values. Defaults to None.

    Returns:
        pd.DataFrame: The preprocessed dataframe.
    """

    # Drop duplicate rows inplace
    df.drop_duplicates(inplace=True)

    # Drop NaN values inplace
    df.dropna(subset=columns_to_check, how="all", inplace=True)

    def extract_year(string):
        match = re.search(r"(\d{4})/(\d{2})", string)
        if match:
            year1, year2 = match.group(1), match.group(2)
            return year1, "20" + year2
        return None, None

    for column in df.select_dtypes(include="category"):
        df[column] = df[column].apply(lambda x: re.sub(r"^\d+\. ", "", str(x)))

    if "wave" in primary_columns:
        # Get the index position of the passed column
        column_pos = df.columns.get_loc("wave")

        df["Wave"] = df["wave"].str.extract(r"Wave (\d+)").astype("int64")
        df[["year1", "year2"]] = df.apply(
            lambda x: extract_year(x["wave"]), axis=1
        ).apply(pd.Series)
        df["year"] = df.apply(
            lambda x: x["year1"] if x["visit"] == "Post-Planting" else x["year2"],
            axis=1,
        )

        # dropping wave, year1 and year2
        df.pop("wave"), df.pop("year1"), df.pop("year2")

        # Convert Year and day columns to numeric type
        df["year"] = pd.to_numeric(df["year"])

        # repositioning
        df.insert(column_pos, "wave", df.pop("Wave"))
        df.insert(column_pos + 1, "year", df.pop("year"))

    if "hhid" in primary_columns:
        df["hhid"] = df["hhid"].astype("int64")

    return df.reset_index(drop=True, inplace=True)


def prepare_concatenated_data(
    filenames_indicator: Dict[str, set], files_paths: Dict[str, str], geodata_path: str
) -> Optional[List[pd.DataFrame]]:
    """
    Prepares concatenated dataframes based on the given filenames and their associated indicators.

    Parameters:
        filenames_indicator (Dict[str, set]): A dictionary where keys are filenames and values are sets of indicators.
        files_paths (Dict[str, str]): A dictionary where keys are filenames and values are file paths.
        geodata_path (str): The file path of the geolocation data.

    Returns:
        List[pd.DataFrame] or None: A list of concatenated dataframes.
    """
    container = defaultdict(list)

    for filename, indicators in filenames_indicator.items():
        mod_filename_match = re.search(r"_mod_[a-z].*", filename).group()

        if mod_filename_match and filename in files_paths:
            file_path = files_paths[filename]
            
            indicators = sorted(list(indicators))
            use_columns = PRIMARY_COLUMNS + indicators

            data = dataframe_reader(file_path, use_columns=use_columns)
            
            preprocessing_transformer(data, PRIMARY_COLUMNS, indicators)
            container[mod_filename_match].append(data)
            
    if container:
        # TODO: make a general reader
        geodata_across_waves = dataframe_reader(file_path=geodata_path, use_columns=['wave', 'hhid', 'longitude', 'latitude'])
        concantenated_data = []

        for data_values in container.values():
            # vertically stacking columns that exists for the same nup_p*_mod_[*]
            data = pd.concat(data_values, axis=0)

            # adding geolocation - longitude and latitude
            data_with_geolocation = pd.merge(
                data, geodata_across_waves, on=["wave", "hhid"], how="inner"
            )
            concantenated_data.append(data_with_geolocation)

        return concantenated_data
    else:
        print("No data to be processed.")
        return None


def indicator_merger(concatenated_data: List[pd.DataFrame]) -> pd.DataFrame:
    """
    Merge and process a list of DataFrames containing indicator data.

    Args:
        concatenated_data (List[pd.DataFrame]): A list of DataFrames to be merged.

    Returns:
        pd.DataFrame: The merged DataFrame after processing.
    """
    data = pd.concat(concatenated_data)
    data.drop_duplicates(inplace=True)
    data.sort_values(
        by=["wave", "year", "zone", "state", "lga", "sector", "ea", "hhid"],
        inplace=True,
    )
    data.dropna(subset=data.columns[9:], how="all", inplace=True)
    data.reset_index(drop=True, inplace=True)
    return data


def harmonize():
    """
    Runs all setup on the harmonize dataset and save for all selected indicators_
    """    
    xml_to_json_converter(XML_FILE_PATH, JSON_FILE_PATH)
    
    data_files_descriptions = prepare_files_descriptions_data(JSON_FILE_PATH)
    write_dict_to_file(
        data_files_descriptions, f"{DATA_FOLDER_PATH}", "data_dictionary"
    )

    # indicator_domain_mapping, columns_filename_lookup, file_columns_question_lookup, filenames_indicator,

    _, _, _, filenames_indicator = add_extract_dict_to_excel(
        DATA_DICTIONARY_PATH, data_files_descriptions, PRIMARY_COLUMNS
    )
    
    harmonized_data = get_all_harmonized_files_dictionary(
        HARMONIZED_DATA_PATH, FILE_EXTENSION_CHOICES
    )
    
    preprocessed_concatenated_data = prepare_concatenated_data(
        filenames_indicator, harmonized_data, GEOLOCATION_DATA_PATH
    )
    
    merged_indicator_data = indicator_merger(preprocessed_concatenated_data)
    saving_data_formatter(merged_indicator_data, "all_domains_combined")
    print(merged_indicator_data.head())